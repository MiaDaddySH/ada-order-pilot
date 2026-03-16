import os
from io import BytesIO
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from app.main import app
from app.repository import OrderRepository
from app.db import init_db

os.environ["PARSE_MODE"] = "fallback"


def test_health() -> None:
    client = TestClient(app)
    response = client.get("/health")
    assert response.status_code == 200
    assert response.json() == {"status": "ok"}


def test_parse_order_input_llm_only_without_key_returns_503(tmp_path: Path) -> None:
    os.environ["DB_PATH"] = str(tmp_path) + "/test_llm_only.db"
    os.environ["PARSE_MODE"] = "llm_only"
    os.environ["LLM_API_KEY"] = ""
    os.environ["LLM_BASE_URL"] = ""
    client = TestClient(app)
    response = client.post("/api/v1/parse-order-input", json={"input_text": "测试地址 王明 13800138000"})
    assert response.status_code == 503
    image_resp = client.post(
        "/api/v1/recipients/import-image",
        json={"image_base64": "aGVsbG8gd29ybGQ=", "mime_type": "image/png"},
    )
    assert image_resp.status_code == 503
    os.environ["PARSE_MODE"] = "fallback"


def test_resolve_product_code_for_short_alias_text(tmp_path: Path) -> None:
    db_path = str(tmp_path / "test_resolve_alias.db")
    init_db(db_path)
    repository = OrderRepository(db_path)
    code = repository.resolve_product_code(
        source_text="狮子牛12+8罐，沈钦雨13867104278，浙江省杭州市萧山区蜀山街道山水苑34-1-501",
        product_name="狮子牛12+8罐",
        brand=None,
        stage=None,
    )
    assert code == "42604770514557"


def test_index_page() -> None:
    client = TestClient(app)
    response = client.get("/")
    assert response.status_code == 200
    assert "text/html" in response.headers.get("content-type", "")
    assert "ADA 智能订单系统" in response.text


def test_parse_order_input(tmp_path: Path) -> None:
    os.environ["DB_PATH"] = str(tmp_path) + "/test_parse.db"
    os.environ["LLM_API_KEY"] = ""
    os.environ["LLM_BASE_URL"] = ""
    client = TestClient(app)
    payload = {
        "input_text": "广东省广州市花都区庙南巷42号嘉汇城西区4栋，游锦平13416101033（HO2 4盒）"
    }
    response = client.post("/api/v1/parse-order-input", json=payload)
    assert response.status_code == 200
    body = response.json()
    assert "recipient" in body
    assert "products" in body
    assert isinstance(body["products"], list)
    assert body["recipient"]["province"] == "广东省"
    assert body["recipient"]["city"] == "广州市"
    assert body["recipient"]["district"] == "花都区"
    assert "庙南巷42号" in body["recipient"]["address_detail"]
    assert body["parse_source"] == "fallback"
    assert body["products"][0]["simple_code"] == "HO2"


def test_parse_order_input_with_alias_brand(tmp_path: Path) -> None:
    os.environ["DB_PATH"] = str(tmp_path) + "/test_parse_alias.db"
    os.environ["LLM_API_KEY"] = ""
    os.environ["LLM_BASE_URL"] = ""
    client = TestClient(app)
    payload = {
        "input_text": "浙江省杭州市萧山区蜀山街道山水苑34-1-501\n沈钦雨 13867104278（狮子牛12+8罐）"
    }
    response = client.post("/api/v1/parse-order-input", json=payload)
    assert response.status_code == 200
    body = response.json()
    assert body["recipient"]["name"] == "沈钦雨"
    assert body["recipient"]["province"] == "浙江省"
    assert body["recipient"]["city"] == "杭州市"
    assert body["recipient"]["district"] == "萧山区"
    assert body["recipient"]["address_detail"] == "蜀山街道山水苑34-1-501"
    assert body["parse_source"] == "fallback"
    assert body["products"][0]["simple_code"] is None
    assert body["products"][0]["quantity"] == 8


def test_parse_order_input_with_plus_symbol_for_lewenzan(tmp_path: Path) -> None:
    os.environ["DB_PATH"] = str(tmp_path) + "/test_parse_plus.db"
    os.environ["LLM_API_KEY"] = ""
    os.environ["LLM_BASE_URL"] = ""
    client = TestClient(app)
    payload = {
        "input_text": "直邮 8罐乐温赞 牛奶 6➕\n\n何建荣 电话15235223296\n地址：山西省大同市平城区绿地璀璨天城1号楼"
    }
    response = client.post("/api/v1/parse-order-input", json=payload)
    assert response.status_code == 200
    body = response.json()
    assert body["products"][0]["simple_code"] is None
    assert body["products"][0]["stage"] == "6+"
    assert body["products"][0]["quantity"] == 8


def test_parse_order_input_with_inline_phone_and_address_label(tmp_path: Path) -> None:
    os.environ["DB_PATH"] = str(tmp_path) + "/test_parse_inline_phone.db"
    os.environ["LLM_API_KEY"] = ""
    os.environ["LLM_BASE_URL"] = ""
    client = TestClient(app)
    payload = {
        "input_text": "刘琳琳：电话 18068655678 江苏省 南通市 如东县 城中街道 泰山路22号碧桂园二期别墅5302"
    }
    response = client.post("/api/v1/parse-order-input", json=payload)
    assert response.status_code == 200
    body = response.json()
    assert body["recipient"]["name"] == "刘琳琳"
    assert body["recipient"]["province"] == "江苏省"
    assert body["recipient"]["city"] == "南通市"
    assert body["recipient"]["district"] == "如东县"
    assert "泰山路22号" in body["recipient"]["address_detail"]


def test_create_order_from_input_idempotent(tmp_path: Path) -> None:
    os.environ["DB_PATH"] = str(tmp_path) + "/test_order.db"
    client = TestClient(app)
    payload = {
        "input_text": "广东省广州市花都区庙南巷42号嘉汇城西区4栋，游锦平13416101033身份证440102199001019876（HO2 4盒）"
    }
    first = client.post("/api/v1/orders/from-input", json=payload)
    second = client.post("/api/v1/orders/from-input", json=payload)
    assert first.status_code == 200
    assert second.status_code == 200
    first_body = first.json()
    second_body = second.json()
    assert first_body["order_created"] is True
    assert second_body["order_created"] is False
    assert first_body["order_no"] == second_body["order_no"]
    assert first_body["parse_result"]["products"][0]["simple_code"] == "HO2"
    assert first_body["parse_result"]["recipient"]["id_card_no"] == "440102199001019876"
    assert first_body["parse_result"]["parse_source"] == "fallback"


def test_create_order_from_input_requires_simple_code(tmp_path: Path) -> None:
    os.environ["DB_PATH"] = str(tmp_path) + "/test_order_2.db"
    client = TestClient(app)
    payload = {"input_text": "上海市静安区南京西路100号，王明13800138000（未知奶粉 2盒）"}
    response = client.post("/api/v1/orders/from-input", json=payload)
    assert response.status_code == 422


def test_product_catalog_create_list_and_disable(tmp_path: Path) -> None:
    os.environ["DB_PATH"] = str(tmp_path) + "/test_products.db"
    client = TestClient(app)
    created = client.post(
        "/api/v1/products",
        json={"product_name": "测试奶粉1段800克", "simple_code": "TEST1"},
    )
    assert created.status_code == 200
    product_id = created.json()["id"]
    queried = client.get("/api/v1/products", params={"keyword": "TEST1"})
    assert queried.status_code == 200
    assert any(item["simple_code"] == "TEST1" for item in queried.json())
    disabled = client.patch(f"/api/v1/products/{product_id}/status", json={"status": 0})
    assert disabled.status_code == 200
    active_only = client.get("/api/v1/products", params={"keyword": "TEST1"})
    assert active_only.status_code == 200
    assert all(item["id"] != product_id for item in active_only.json())
    include_inactive = client.get(
        "/api/v1/products",
        params={"keyword": "TEST1", "include_inactive": True},
    )
    assert include_inactive.status_code == 200
    assert any(item["id"] == product_id and item["status"] == 0 for item in include_inactive.json())


def test_product_batch_upsert(tmp_path: Path) -> None:
    os.environ["DB_PATH"] = str(tmp_path) + "/test_products_batch.db"
    client = TestClient(app)
    response = client.post(
        "/api/v1/products/batch-upsert",
        json={
            "products": [
                {"product_name": "批量奶粉A", "simple_code": "BATCHA"},
                {"product_name": "批量奶粉B", "simple_code": "BATCHB"},
            ]
        },
    )
    assert response.status_code == 200
    assert response.json()["upserted_count"] == 2
    queried = client.get("/api/v1/products", params={"keyword": "BATCH"})
    assert queried.status_code == 200
    codes = {item["simple_code"] for item in queried.json()}
    assert "BATCHA" in codes
    assert "BATCHB" in codes


def test_export_templates(tmp_path: Path) -> None:
    recipient_template = tmp_path / "收件人模板.xlsx"
    order_template = tmp_path / "订单导入模板.xlsx"
    _create_recipient_template(recipient_template)
    _create_order_template(order_template)
    os.environ["DB_PATH"] = str(tmp_path / "test_export.db")
    os.environ["RECIPIENT_TEMPLATE_PATH"] = str(recipient_template)
    os.environ["ORDER_TEMPLATE_PATH"] = str(order_template)
    os.environ["EXPORT_DIR"] = str(tmp_path / "exports")
    client = TestClient(app)
    payload = {
        "input_text": "广东省广州市花都区庙南巷42号嘉汇城西区4栋，游锦平13416101033身份证440102199001019876（HO2 4盒）"
    }
    created = client.post("/api/v1/orders/from-input", json=payload)
    assert created.status_code == 200
    recipients_file = client.get("/api/v1/export/recipients-template")
    orders_file = client.get("/api/v1/export/orders-template", params={"recent_days": 7, "limit": 10})
    orders_data_file = client.get("/api/v1/export/orders", params={"recent_days": 7, "limit": 10})
    assert recipients_file.status_code == 200
    assert orders_file.status_code == 200
    assert orders_data_file.status_code == 200
    assert recipients_file.content[:2] == b"PK"
    assert orders_file.content[:2] == b"PK"
    assert orders_data_file.content[:2] == b"PK"
    workbook = load_workbook(BytesIO(orders_data_file.content))
    worksheet = workbook[workbook.sheetnames[0]]
    assert worksheet.cell(row=2, column=9).value not in (None, "")


def test_management_crud_apis(tmp_path: Path) -> None:
    os.environ["DB_PATH"] = str(tmp_path / "test_manage.db")
    client = TestClient(app)

    created_recipient = client.post(
        "/api/v1/recipients",
        json={
            "name": "张三",
            "phone": "13800001111",
            "province": "浙江省",
            "city": "杭州市",
            "district": "萧山区",
            "address_detail": "测试路1号",
            "raw_address": "浙江省杭州市萧山区测试路1号",
            "postcode": "310000",
        },
    )
    assert created_recipient.status_code == 200
    recipient_id = created_recipient.json()["id"]
    updated_recipient = client.put(
        f"/api/v1/recipients/{recipient_id}",
        json={
            "name": "李四",
            "phone": "13800002222",
            "province": "浙江省",
            "city": "杭州市",
            "district": "滨江区",
            "address_detail": "更新路2号",
            "raw_address": "浙江省杭州市滨江区更新路2号",
            "postcode": "310051",
        },
    )
    assert updated_recipient.status_code == 200
    assert updated_recipient.json()["name"] == "李四"
    recipient_batch = client.post(
        "/api/v1/recipients/batch-upsert",
        json={
            "recipients": [
                {
                    "name": "批量收件人A",
                    "phone": "13800003333",
                    "id_card_no": "320102199001019876",
                    "province": "江苏省",
                    "city": "南京市",
                    "district": "鼓楼区",
                    "address_detail": "中山路100号",
                    "raw_address": "江苏省南京市鼓楼区中山路100号",
                    "postcode": "210000",
                }
            ]
        },
    )
    assert recipient_batch.status_code == 200
    assert recipient_batch.json()["imported_count"] == 1

    created_order = client.post(
        "/api/v1/orders",
        json={
            "recipient_id": recipient_id,
            "source_text": "手工创建",
            "confidence": 1.0,
            "needs_review": False,
            "status": "ready_to_upload",
            "items": [
                {
                    "simple_code": "HO2",
                    "brand": "Holle",
                    "product_name": "Holle有机婴幼儿牛奶粉2段600g",
                    "stage": "2段",
                    "quantity": 2,
                    "unit": "盒",
                }
            ],
        },
    )
    assert created_order.status_code == 200
    order_id = created_order.json()["id"]
    order_updated = client.put(f"/api/v1/orders/{order_id}", json={"status": "pending_review", "needs_review": True})
    assert order_updated.status_code == 200
    assert order_updated.json()["status"] == "pending_review"

    product_created = client.post("/api/v1/products", json={"product_name": "测试商品X", "simple_code": "TXX"})
    assert product_created.status_code == 200
    product_id = product_created.json()["id"]
    product_updated = client.put(f"/api/v1/products/{product_id}", json={"status": 0})
    assert product_updated.status_code == 200
    assert product_updated.json()["status"] == 0

    senders = client.get("/api/v1/senders")
    assert senders.status_code == 200
    sender_created = client.post(
        "/api/v1/senders",
        json={
            "name": "新寄件人",
            "phone": "13900003333",
            "street": "Main St",
            "house_no": "8A",
            "postcode": "70100",
            "city": "Stuttgart",
            "country_code": "DE",
            "is_default": True,
        },
    )
    assert sender_created.status_code == 200
    sender_id = sender_created.json()["id"]
    sender_updated = client.put(
        f"/api/v1/senders/{sender_id}",
        json={
            "name": "寄件人B",
            "phone": "13900004444",
            "street": "Second St",
            "house_no": "9",
            "postcode": "70200",
            "city": "Berlin",
            "country_code": "DE",
            "is_default": True,
        },
    )
    assert sender_updated.status_code == 200
    assert sender_updated.json()["name"] == "寄件人B"
    sender_batch = client.post(
        "/api/v1/senders/batch-upsert",
        json={
            "senders": [
                {
                    "name": "批量寄件人A",
                    "phone": "13900005555",
                    "street": "Third St",
                    "house_no": "10",
                    "postcode": "70300",
                    "city": "Munich",
                    "country_code": "DE",
                    "is_default": False,
                }
            ]
        },
    )
    assert sender_batch.status_code == 200
    assert sender_batch.json()["imported_count"] == 1

    assert client.delete(f"/api/v1/orders/{order_id}").status_code == 200
    assert client.delete(f"/api/v1/recipients/{recipient_id}").status_code == 200
    assert client.delete(f"/api/v1/products/{product_id}").status_code == 200
    assert client.delete(f"/api/v1/senders/{sender_id}").status_code == 200


def _create_recipient_template(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "导入数据"
    sheet["A1"] = "中国地址模板导入"
    headers = ["*姓名", "*身份证号码", "*电话国际区号", "*电话号码", "*省", "*市", "*区", "*详细地址", "*邮编"]
    for idx, header in enumerate(headers, start=1):
        sheet.cell(row=2, column=idx).value = header
    sheet["C3"] = "86"
    workbook.save(path)


def _create_order_template(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "批量下单"
    headers = [
        "包裹备注",
        "寄件人姓名",
        "寄件人电话",
        "路名",
        "门牌号",
        "寄件人邮编",
        "寄件人城市",
        "寄件人国家简称",
        "收件人姓名",
        "身份证号",
        "手机号码",
        "收件人国家简称",
        "省",
        "市",
        "区/县",
        "详细地址（省市区/县请勿重复填）",
        "渠道代码",
        "货物用途",
        "商品代码1",
        "数量1",
        "商品代码2",
        "数量2",
        "商品代码3",
        "数量3",
        "商品代码4",
        "数量4",
        "商品代码5",
        "数量5",
        "商品代码6",
        "数量6",
    ]
    for idx, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=idx).value = header
    workbook.save(path)
