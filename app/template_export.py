from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.llm_client import Settings


class TemplateExporter:
    def __init__(self, settings: Settings) -> None:
        self.settings = settings

    def export_recipients(self, recipients: list[dict[str, object]]) -> Path:
        return self._export_with_template(
            template_path=Path(self.settings.recipient_template_path),
            output_prefix="收件人导出",
            rows=self._build_recipient_rows(recipients),
        )

    def export_orders(self, orders: list[dict[str, object]], sender_profile: dict[str, object] | None = None) -> Path:
        return self._export_with_template(
            template_path=Path(self.settings.order_template_path),
            output_prefix="订单导出",
            rows=self._build_order_rows(orders, sender_profile=sender_profile),
        )

    def _export_with_template(
        self,
        template_path: Path,
        output_prefix: str,
        rows: list[dict[str, object]],
    ) -> Path:
        workbook = load_workbook(template_path)
        worksheet = workbook[workbook.sheetnames[0]]
        header_row_index = self._detect_header_row(worksheet)
        write_row = header_row_index + 1
        if write_row <= worksheet.max_row:
            self._clear_rows(worksheet, write_row)
        headers = self._headers(worksheet, header_row_index)
        for row_data in rows:
            for col, header in headers.items():
                worksheet.cell(write_row, col).value = row_data.get(header)
            write_row += 1
        export_dir = Path(self.settings.export_dir)
        export_dir.mkdir(parents=True, exist_ok=True)
        output_path = export_dir / f"{output_prefix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        workbook.save(output_path)
        return output_path

    def _detect_header_row(self, worksheet: Worksheet) -> int:
        best_row = 1
        best_score = 0
        for row_index in range(1, min(worksheet.max_row, 20) + 1):
            values = [worksheet.cell(row_index, col).value for col in range(1, worksheet.max_column + 1)]
            non_empty = [str(value).strip() for value in values if value is not None and str(value).strip()]
            score = len(non_empty)
            if score > best_score:
                best_score = score
                best_row = row_index
        return best_row

    def _headers(self, worksheet: Worksheet, row_index: int) -> dict[int, str]:
        headers: dict[int, str] = {}
        for col in range(1, worksheet.max_column + 1):
            value = worksheet.cell(row_index, col).value
            if value is None:
                continue
            key = str(value).strip()
            if key:
                headers[col] = key
        return headers

    def _clear_rows(self, worksheet: Worksheet, start_row: int) -> None:
        for row in range(start_row, worksheet.max_row + 1):
            for col in range(1, worksheet.max_column + 1):
                worksheet.cell(row, col).value = None

    def _build_recipient_rows(self, recipients: list[dict[str, object]]) -> list[dict[str, object]]:
        rows: list[dict[str, object]] = []
        for item in recipients:
            rows.append(
                {
                    "*姓名": item.get("name"),
                    "*身份证号码": "",
                    "*电话国际区号": "86",
                    "*电话号码": item.get("phone"),
                    "*省": item.get("province"),
                    "*市": item.get("city"),
                    "*区": item.get("district"),
                    "*详细地址": item.get("address_detail"),
                    "*邮编": item.get("postcode"),
                }
            )
        return rows

    def _build_order_rows(
        self,
        orders: list[dict[str, object]],
        sender_profile: dict[str, object] | None = None,
    ) -> list[dict[str, object]]:
        rows: list[dict[str, object]] = []
        sender_name = str(sender_profile["name"]) if sender_profile is not None else self.settings.sender_name
        sender_phone = str(sender_profile["phone"]) if sender_profile is not None else self.settings.sender_phone
        sender_street = str(sender_profile["street"]) if sender_profile is not None else self.settings.sender_street
        sender_house_no = str(sender_profile["house_no"]) if sender_profile is not None else self.settings.sender_house_no
        sender_postcode = str(sender_profile["postcode"]) if sender_profile is not None else self.settings.sender_postcode
        sender_city = str(sender_profile["city"]) if sender_profile is not None else self.settings.sender_city
        sender_country_code = (
            str(sender_profile["country_code"]) if sender_profile is not None else self.settings.sender_country_code
        )
        for order in orders:
            row: dict[str, Any] = {
                "包裹备注": "",
                "寄件人姓名": sender_name,
                "寄件人电话": sender_phone,
                "路名": sender_street,
                "门牌号": sender_house_no,
                "寄件人邮编": sender_postcode,
                "寄件人城市": sender_city,
                "寄件人国家简称": sender_country_code,
                "收件人姓名": order.get("recipient_name"),
                "身份证号": "",
                "手机号码": order.get("recipient_phone"),
                "收件人国家简称": self.settings.recipient_country_code,
                "省": order.get("province"),
                "市": order.get("city"),
                "区/县": order.get("district"),
                "详细地址（省市区/县请勿重复填）": order.get("address_detail"),
                "渠道代码": self.settings.channel_code,
                "货物用途": self.settings.goods_purpose,
            }
            raw_items = order.get("items", [])
            items: list[dict[str, object]] = []
            if isinstance(raw_items, list):
                for item in raw_items:
                    if isinstance(item, dict):
                        items.append(item)
            for idx in range(1, 7):
                code_key = f"商品代码{idx}"
                qty_key = f"数量{idx}"
                if idx <= len(items):
                    current = items[idx - 1]
                    row[code_key] = current.get("simple_code")
                    row[qty_key] = current.get("quantity")
                else:
                    row[code_key] = ""
                    row[qty_key] = ""
            rows.append(row)
        return rows
